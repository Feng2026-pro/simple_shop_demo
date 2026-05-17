#!/usr/bin/env python3
"""
Convert scenario-review Excel into executable API cases JSON.

场景 Excel 可由 `scenario_json_to_excel.py` 从 `*_scenario.json` 生成（见 references/scenario-json-schema.md）。

固定列（与 scenario-excel-spec 一致）：
- 场景ID、场景名称、业务目标、前置条件、接口路径、请求方法
- 关键输入（自然语言）：JSON:{...}；GET/DELETE 可空，等价 {}
- 预期结果（自然语言）：CHECK:...
- 风险等级（P0/P1/P2）、是否转执行用例（Y/N）、备注

可选列：
- 预期状态码 / 预期状态码（HTTP）：默认 200
- 路径参数：JSON 字符串，替换路径 {key}
- 前置依赖、依赖产出提取、注入方式：与 runner Excel 列一致
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

REQUIRED_COLUMNS = [
    "场景ID",
    "场景名称",
    "接口路径",
    "请求方法",
    "关键输入（自然语言）",
    "预期结果（自然语言）",
    "风险等级（P0/P1/P2）",
    "是否转执行用例（Y/N）",
]

METHODS = {"GET", "POST", "PUT", "DELETE"}


def _read_rows(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        wb.close()
        return []
    keys = [str(c).lstrip("\ufeff").strip() if c is not None else "" for c in header]
    out = []
    for row in rows_iter:
        one = {}
        for i, k in enumerate(keys):
            if not k:
                continue
            v = row[i] if i < len(row) else None
            one[k] = "" if v is None else str(v).strip()
        if any(one.values()):
            out.append(one)
    wb.close()
    return out


def _parse_json_input(raw: str, method: str):
    txt = (raw or "").strip()
    # GET/DELETE 无请求体时允许留空或写「无」，等价 JSON:{}
    if not txt or txt in ("-", "无", "无请求体"):
        if method in ("GET", "DELETE"):
            return "{}", None
        return None, "关键输入未按 JSON:{...} 格式填写"
    if not txt.startswith("JSON:"):
        return None, "关键输入未按 JSON:{...} 格式填写"
    body_str = txt[5:].strip()
    if not body_str:
        if method in ("GET", "DELETE"):
            return "{}", None
        return None, "关键输入 JSON 内容为空"
    try:
        obj = json.loads(body_str)
    except json.JSONDecodeError:
        return None, "关键输入 JSON 解析失败"
    if not isinstance(obj, (dict, list)):
        return None, "关键输入 JSON 必须是对象或数组"
    # runner 接收字符串形式的 JSON
    return json.dumps(obj, ensure_ascii=False), None


def _parse_check(raw: str):
    txt = (raw or "").strip()
    if not txt.startswith("CHECK:"):
        return None, "预期结果未按 CHECK:... 格式填写"
    expr = txt[6:].strip()
    if not expr:
        return None, "预期结果 CHECK 表达式为空"
    return expr, None


def _build_case(row: dict):
    sid = (row.get("场景ID") or "").strip()
    name = (row.get("场景名称") or "").strip()
    path = (row.get("接口路径") or "").strip()
    method = (row.get("请求方法") or "").strip().upper()
    risk = (row.get("风险等级（P0/P1/P2）") or "P1").strip().upper()
    to_run = (row.get("是否转执行用例（Y/N）") or "N").strip().upper()

    if to_run != "Y":
        return None, "是否转执行用例 != Y"
    if not sid or not name or not path or not method:
        return None, "缺少必填字段（场景ID/场景名称/接口路径/请求方法）"
    if method not in METHODS:
        return None, f"请求方法非法: {method}"
    if risk not in {"P0", "P1", "P2"}:
        risk = "P1"

    req_body, err = _parse_json_input(row.get("关键输入（自然语言）", ""), method)
    if err:
        return None, err
    check_expr, err = _parse_check(row.get("预期结果（自然语言）", ""))
    if err:
        return None, err

    status_raw = (row.get("预期状态码") or row.get("预期状态码（HTTP）") or "200").strip()
    try:
        expected_http = int(status_raw)
    except ValueError:
        return None, f"预期状态码非法: {status_raw!r}"

    path_params = (row.get("路径参数") or "").strip()
    if path_params:
        try:
            json.loads(path_params)
        except json.JSONDecodeError:
            return None, "路径参数须为合法 JSON 对象字符串"

    dep = (row.get("前置依赖") or "无").strip() or "无"
    extract = (row.get("依赖产出提取") or "").strip()
    inject = (row.get("注入方式") or "").strip()

    one = {
        "用例ID": sid,
        "用例名称": name,
        "接口路径": path,
        "路径参数": path_params,
        "请求方法": method,
        "请求体/参数": req_body,
        "预期状态码": expected_http,
        "预期响应校验": check_expr,
        "优先级": risk,
        "前置依赖": dep,
        "依赖产出提取": extract,
        "注入方式": inject,
        "是否运行": "是",
        "运行结果": "",
    }
    return one, None


def main():
    parser = argparse.ArgumentParser(description="Convert scenario Excel to executable cases JSON")
    parser.add_argument("-i", "--input", required=True, help="Scenario Excel path (.xlsx)")
    parser.add_argument("-o", "--output", required=True, help="Output cases JSON path")
    parser.add_argument(
        "--issues",
        default=None,
        help="Output mapping issues markdown path (default: <output_stem>_mapping_issues.md)",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.input)
    out_path = Path(args.output)
    issues_path = Path(args.issues) if args.issues else out_path.with_name(f"{out_path.stem}_mapping_issues.md")

    rows = _read_rows(xlsx_path)
    if not rows:
        raise SystemExit(f"No rows found in: {xlsx_path}")

    missing_cols = [c for c in REQUIRED_COLUMNS if c not in rows[0]]
    if missing_cols:
        raise SystemExit(f"Missing required columns: {', '.join(missing_cols)}")

    cases = []
    issues = []

    for row in rows:
        sid = (row.get("场景ID") or "").strip() or "<NO_ID>"
        sname = (row.get("场景名称") or "").strip() or "<NO_NAME>"
        one, err = _build_case(row)
        if err:
            issues.append(f"- {sid} {sname}：{err}")
            continue
        cases.append(one)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps({"cases": cases}, ensure_ascii=False, indent=2), encoding="utf-8")

    issues_path.parent.mkdir(parents=True, exist_ok=True)
    md = [
        "# 场景映射问题清单",
        "",
        f"- 输入行数：{len(rows)}",
        f"- 成功映射：{len(cases)}",
        f"- 待补充：{len(issues)}",
        "",
    ]
    if issues:
        md.extend(issues)
    else:
        md.append("- 无")
    issues_path.write_text("\n".join(md) + "\n", encoding="utf-8")

    print(f"Generated cases JSON: {out_path}")
    print(f"Generated mapping issues: {issues_path}")


if __name__ == "__main__":
    main()
