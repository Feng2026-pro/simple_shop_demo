#!/usr/bin/env python3
"""
将用例 JSON 转为 Excel 测试用例表。用例设计（正常+异常）由大模型根据接口文档产出 JSON，本脚本仅做转换。

Usage:
    doc_to_excel.py --input <cases.json> --output <excel_path>
    doc_to_excel.py -i cases.json -o test_cases.xlsx
"""

import argparse
import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None
    Font = None


# Excel 列顺序（与 excel-column-spec.md、doc_to_csv.py 一致）
COLUMNS = [
    "用例ID", "用例名称", "接口路径", "路径参数", "请求方法", "请求体/参数",
    "预期状态码", "预期响应校验", "优先级", "前置依赖", "依赖产出提取", "注入方式",
    "是否运行", "运行结果",
]


def _require_openpyxl():
    if Workbook is None:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")


def load_cases_from_json(file_path: Path) -> list[dict]:
    """
    从 JSON 文件加载用例列表。
    支持根节点为数组，或根节点为对象且含 "cases" 键的数组。
    每条用例规范化为包含所有 COLUMNS 键的 dict，缺失键填空字符串。
    """
    raw = file_path.read_text(encoding="utf-8")
    data = json.loads(raw)

    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict) and "cases" in data:
        rows = data["cases"]
    else:
        raise ValueError("JSON root must be an array or an object with key 'cases' (array)")

    if not isinstance(rows, list):
        raise ValueError("Cases must be a list")

    cases = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        case = {}
        for col in COLUMNS:
            val = item.get(col, "")
            if val is None:
                val = ""
            case[col] = val
        # 默认「是否运行」为「是」，「运行结果」为空
        if case.get("是否运行") == "":
            case["是否运行"] = "是"
        if case.get("运行结果") is None:
            case["运行结果"] = ""
        # 预期状态码：数字保留，字符串尝试转 int
        status = case.get("预期状态码", 200)
        if isinstance(status, str) and status.strip():
            try:
                case["预期状态码"] = int(status.strip())
            except ValueError:
                case["预期状态码"] = 200
        elif not isinstance(status, int):
            case["预期状态码"] = 200
        cases.append(case)

    return cases


def write_excel(cases: list[dict], output_path: Path) -> None:
    """Write test cases to Excel file."""
    _require_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "用例"

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        if Font:
            cell.font = Font(bold=True)

    for row_idx, case in enumerate(cases, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=case.get(col_name, ""))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Convert test cases JSON to Excel (column format per excel-column-spec)"
    )
    parser.add_argument("--input", "-i", required=True, help="Input JSON file path (array or { \"cases\": [...] })")
    parser.add_argument("--output", "-o", required=True, help="Output Excel file path")
    args = parser.parse_args()

    inp = Path(args.input)
    out = Path(args.output)
    if not inp.exists():
        raise FileNotFoundError(f"Input file not found: {inp}")

    cases = load_cases_from_json(inp)
    write_excel(cases, out)
    print(f"Generated {len(cases)} test cases to {out}")


if __name__ == "__main__":
    main()
